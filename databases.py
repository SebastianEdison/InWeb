import sqlite3
import hashlib
import pytz
import sys
import os
from datetime import datetime

tz_chile = pytz.timezone('America/Santiago')
fecha_chile = datetime.now(tz_chile).strftime('%Y-%m-%d %H:%M:%S')

def get_app_dir():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

def conectar():
    db_path = os.path.join(get_app_dir(), 'inventario.db')
    conexion = sqlite3.connect(db_path)
    conexion.row_factory = sqlite3.Row
    return conexion

def crear_tablas():

    conexion = conectar()
    print("Base de datos conectada")
    cursor = conexion.cursor()
    print("Cursor creado")

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS productos (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    codigo_barra TEXT UNIQUE,
                    nombre TEXT NOT NULL,
                    precio_venta REAL NOT NULL,
                    costo REAL,
                    stock INTEGER DEFAULT 0 CHECK(stock >= 0) ,
                    activo INTEGER DEFAULT 1,
                    unidad TEXT DEFAULT 'Unidad'
                    );
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS ventas (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    fecha DATETIME DEFAULT (strftime('%Y-%m-%d %H:%M:%S', 'now', '-3 hours')),
                    total REAL NOT NULL,
                    metodo_pago TEXT
                    );   
    """)

    cursor.execute("""                                
        CREATE TABLE IF NOT EXISTS detalle_venta(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                venta_id INTEGER NOT NULL,
                producto_id INTEGER NOT NULL,
                cantidad INTEGER NOT NULL,
                precio_unitario REAL NOT NULL,
                subtotal REAL NOT NULL,
                FOREIGN KEY (venta_id)REFERENCES ventas(id),
                FOREIGN KEY (producto_id) REFERENCES productos(id)
                );
    """)
                
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS movimientos_stock(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                producto_id INTEGER,
                tipo TEXT,
                cantidad INTEGER,
                fecha DATETIME DEFAULT (strftime('%Y-%m-%d %H:%M:%S', 'now', '-3 hours')),
                FOREIGN KEY (producto_id) REFERENCES productos(id)
                );
    """)
    
    # NUEVA TABLA DE CIERRES
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS cierres (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha TEXT,
            efectivo INTEGER,
            tarjeta INTEGER,
            otros INTEGER,
            fiados INTEGER,
            total INTEGER,
            turno TEXT
        )
    ''')
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS usuarios (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL,
        nombre TEXT NOT NULL,
        rol TEXT DEFAULT 'empleado',
        activo INTEGER DEFAULT 1
    )
    ''')
    # Usuario admin por defecto (solo si no existe)
    cursor.execute("SELECT id FROM usuarios WHERE username = 'admin'")
    if not cursor.fetchone():
            import hashlib
            password_hash = hashlib.sha256('admin123'.encode()).hexdigest()
            cursor.execute("""
            INSERT INTO usuarios (username, password, nombre, rol)
                VALUES ('admin', ?, 'Administrador', 'admin')
            """, (password_hash,))
            print("✅ Usuario admin creado: admin / admin123")

    cursor.execute('''
    CREATE TABLE IF NOT EXISTS fiados (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre_cliente TEXT NOT NULL,
        monto_total REAL NOT NULL,
        monto_pagado REAL DEFAULT 0,
        fecha TEXT NOT NULL,
        estado TEXT DEFAULT 'pendiente',
        detalle TEXT
    )
    ''')
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS facturas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        numero_factura TEXT NOT NULL,
        proveedor TEXT NOT NULL,
        rut_proveedor TEXT,
        fecha TEXT NOT NULL,
        monto_total REAL NOT NULL,
        productos TEXT,
        estado TEXT DEFAULT 'pendiente'
    )
''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS configuracion (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            clave TEXT UNIQUE NOT NULL,
            valor TEXT
        )
    ''')

    # Configuración por defecto
    configs = [
        ('nombre_negocio', 'EL HISTORICO'),
        ('rut_negocio', ''),
        ('aplica_iva', '1'),
    ]
    for clave, valor in configs:
        cursor.execute("""
            INSERT OR IGNORE INTO configuracion (clave, valor) 
            VALUES (?, ?)
        """, (clave, valor))
    try:
        cursor.execute("ALTER TABLE productos ADD COLUMN fecha_vencimiento TEXT DEFAULT NULL")
    except:
        pass

    # Feature 1: stock mínimo por producto
    try:
        cursor.execute("ALTER TABLE productos ADD COLUMN stock_minimo INTEGER DEFAULT 0")
    except:
        pass

    # Feature 2: categoría de producto
    try:
        cursor.execute("ALTER TABLE productos ADD COLUMN categoria TEXT DEFAULT 'General'")
    except:
        pass

    # Feature 3: descuento en ventas
    try:
        cursor.execute("ALTER TABLE ventas ADD COLUMN descuento REAL DEFAULT 0")
    except:
        pass

    # Feature 5: anulación de ventas
    try:
        cursor.execute("ALTER TABLE ventas ADD COLUMN anulada INTEGER DEFAULT 0")
    except:
        pass

    # Feature 6: motivo en movimientos de stock
    try:
        cursor.execute("ALTER TABLE movimientos_stock ADD COLUMN motivo TEXT")
    except:
        pass

    # Feature 7: tabla de proveedores
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS proveedores (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            rut TEXT,
            telefono TEXT,
            email TEXT,
            notas TEXT,
            activo INTEGER DEFAULT 1
        )
    ''')

    conexion.commit()
    conexion.close()

def agregar_producto(codigo, nombre, precio, costo, stock, unidad='Unidad', fecha_vencimiento=None, stock_minimo=0, categoria='General'):
    conexion = conectar()
    conexion.row_factory = sqlite3.Row
    cursor = conexion.cursor()

    try:
        cursor.execute("SELECT id, stock FROM productos WHERE codigo_barra = ?", (codigo,))
        existente = cursor.fetchone()

        if existente:
            id_producto = existente['id']
            stock_actual = existente['stock']
            nuevo_total = stock_actual + int(stock)

            cursor.execute("""
                UPDATE productos
                SET nombre=?, precio_venta=?, costo=?, stock=?, unidad=?, fecha_vencimiento=?, stock_minimo=?, categoria=?
                WHERE id=?
            """, (nombre, precio, costo, nuevo_total, unidad, fecha_vencimiento, stock_minimo, categoria, id_producto))
            print(f"✅ Producto '{nombre}' actualizado. Nuevo stock: {nuevo_total}")

        else:
            cursor.execute("""
                INSERT INTO productos (codigo_barra, nombre, precio_venta, costo, stock, unidad, fecha_vencimiento, stock_minimo, categoria)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (codigo, nombre, precio, costo, stock, unidad, fecha_vencimiento, stock_minimo, categoria))
            print(f"✨ Nuevo producto '{nombre}' registrado")

        conexion.commit()

    except sqlite3.Error as e:
        conexion.rollback()
        print(f"❌ Error al procesar producto: {e}")
    finally:
        conexion.close()
            
def obtener_productos(nombre_buscar=None, categoria=None):
    conexion = conectar()
    conexion.row_factory = sqlite3.Row
    cursor = conexion.cursor()

    try:
        condiciones = []
        params = []

        if nombre_buscar:
            condiciones.append("(nombre LIKE ? OR codigo_barra LIKE ?)")
            param = f"%{nombre_buscar}%"
            params.extend([param, param])

        if categoria:
            condiciones.append("categoria = ?")
            params.append(categoria)

        sql = "SELECT * FROM productos"
        if condiciones:
            sql += " WHERE " + " AND ".join(condiciones)

        cursor.execute(sql, params)
        resultados = cursor.fetchall()
        return resultados

    except sqlite3.Error as e:
        print(f"Error en la consulta: {e}")
        return []

    finally:
        if conexion:
            conexion.close()

def buscar_producto_por_codigo(codigo):
    conexion = None
    try:
        conexion = conectar()
        print("Base de datos conectada")
        cursor = conexion.cursor()
        #se ejecuta la consulta
        cursor.execute(
            "SELECT * FROM productos WHERE codigo_barra = ?",
            (codigo,)
        )
        #fetchone() devuelve una tupla si existe, o None si no existe
        producto = cursor.fetchone()

        return producto    
    except sqlite3.Error as e:
        print(f"Error al buscar el producto: {e}")
        return None # Devolvemos None para indicar que hubo un fallo
    
    finally:
        if conexion:
            conexion.close()

def modificar_stock(producto_id, cantidad_cambio, motivo=None):
    conexion = None
    try:
        conexion = conectar()
        cursor = conexion.cursor()

        # verificamos stock actual y nombre
        cursor.execute("SELECT stock, nombre FROM productos WHERE id = ?", (producto_id,))
        resultado = cursor.fetchone()

        if not resultado:
            return False, "Producto no existe"

        stock_actual = resultado[0]
        nombre_prod = resultado[1]
        nuevo_stock = stock_actual + cantidad_cambio

        if nuevo_stock < 0:
            return False, f"Error: Solo hay {stock_actual} unidades de {nombre_prod}."

        cursor.execute("UPDATE productos SET stock = ? WHERE id = ?", (nuevo_stock, producto_id))

        tipo_mov = "ENTRADA" if cantidad_cambio > 0 else "VENTA"
        cursor.execute("""
            INSERT INTO movimientos_stock (producto_id, tipo, cantidad, motivo)
            VALUES (?, ?, ?, ?)
        """, (producto_id, tipo_mov, abs(cantidad_cambio), motivo))

        conexion.commit()
        print(f"Stock de {nombre_prod} actualizado: {nuevo_stock} unidades.")
        return True, "Éxito"

    except sqlite3.Error as e:
        if conexion: conexion.rollback()
        return False, f"Error DB: {e}"
    finally:
        if conexion: conexion.close()

def registrar_venta(carrito, metodo_pago="Efectivo", forzar=False, descuento=0):
    conexion = None
    try:
        conexion = conectar()
        cursor = conexion.cursor()
        if not forzar:
            for item in carrito:
                cursor.execute("SELECT stock, nombre FROM productos WHERE id = ?", (item['id'],))
                producto = cursor.fetchone()
                
                if not producto:
                    return False, f"El producto con ID {item['id']} no existe."
                
                if producto['stock'] < item['cantidad']:
                    return False, f"Stock insuficiente para {producto['nombre']}. Solo quedan {producto['stock']}."

        # 1. VALIDACIÓN: Revisar si hay stock para TODO el carrito antes de empezar
        for item in carrito:
            cursor.execute("SELECT stock, nombre FROM productos WHERE id = ?", (item['id'],))
            producto = cursor.fetchone()
            
            if not producto:
                return False, f"El producto con ID {item['id']} no existe."
            
            # Si lo que quiere vender es mayor a lo que hay, cancelamos
            if producto['stock'] < item['cantidad']:
                return False, f"Stock insuficiente para {producto['nombre']}. Solo quedan {producto['stock']}."

        # 2. CALCULAR TOTAL
        total_venta = sum(item['cantidad'] * item['precio'] for item in carrito)

        # 3. INSERTAR CABECERA (Ventas)
        tz_chile = pytz.timezone('America/Santiago')
        fecha_chile = datetime.now(tz_chile).strftime('%Y-%m-%d %H:%M:%S')
        cursor.execute("INSERT INTO ventas (total, metodo_pago, fecha, descuento) VALUES (?, ?, ?, ?)",
                (total_venta, metodo_pago, fecha_chile, descuento))
        venta_id = cursor.lastrowid 

        # 4. PROCESAR PRODUCTOS
        for item in carrito:
            subtotal = item['cantidad'] * item['precio']

            # Detalle de venta
            cursor.execute("""
                INSERT INTO detalle_venta (venta_id, producto_id, cantidad, precio_unitario, subtotal)
                VALUES (?, ?, ?, ?, ?)
            """, (venta_id, item['id'], item['cantidad'], item['precio'], subtotal))

            # DESCUENTO DE STOCK (Aquí es donde se hacía el negativo antes)
            cursor.execute("UPDATE productos SET stock = stock - ? WHERE id = ?",
                           (item['cantidad'], item['id']))
            
            # Registro en Kardex (Movimientos)
            cursor.execute("""
                INSERT INTO movimientos_stock (producto_id, tipo, cantidad, motivo)
                VALUES (?, 'VENTA', ?, NULL)
            """, (item['id'], item['cantidad']))

        # 5. COMMIT FINAL (Solo se guarda si nada falló arriba)
        conexion.commit()
        print(f"✅ Venta #{venta_id} registrada con éxito (${total_venta})")
        return True, venta_id

    except Exception as e:
        if conexion:
            conexion.rollback() # Si hay error de sistema, deshace todo
        print(f"❌ Error al registrar la venta: {e}")
        return False, str(e)
        
    finally: 
        if conexion:
            conexion.close()

def probar_sistema():
    print("--- INICIANDO PRUEBAS DEL SISTEMA ---")
    crear_tablas()

    #Se agregan productos que no existen
    print("\n[1] Preparando productos...")
    agregar_producto('780001', 'Leche entera 1L', 1500, 1000, 50)
    agregar_producto('780002', 'Pan de molde', 2200, 1500, 20)

    # 2. Buscar un producto para obtener su ID real
    prod = buscar_producto_por_codigo('780001')
    prod2 = buscar_producto_por_codigo('780002')

    if prod and prod2:
        # probamos la funcion de modificar stock
        print(f"\n[2] Reponiendo 10 unidades de {prod[2]}...")
        modificar_stock(prod[0], 10) # Suma 10

        # Probar REGISTRAR VENTA (Carrito)
        print("\n[3] Simulando una venta de carrito...")
        mi_carrito = [
            {'id': prod[0], 'cantidad': 2, 'precio': prod[3]},  # 2 Leches
            {'id': prod2[0], 'cantidad': 1, 'precio': prod2[3]} # 1 Pan
        ]
        
        # Probamos venta con Tarjeta (para ver que guarde el método)
        exito, venta_id = registrar_venta(mi_carrito, "Tarjeta")

        if exito:
            print(f"¡Venta #{venta_id} completada!")
            # Verificar stock final de la leche
            final = buscar_producto_por_codigo('780001')
            print(f"Stock final de {final[2]}: {final[5]} (Empezó en 50, +10 reposición, -2 venta = 58)")

    print("\n --- PRUEBAS FINALIZADAS ---")

def actualizar_producto(id_p, nombre, precio, costo, stock, unidad='Unidad', fecha_vencimiento=None, stock_minimo=0, categoria='General'):
    conexion = conectar()
    cursor = conexion.cursor()
    try:
        cursor.execute("""
            UPDATE productos
            SET nombre=?, precio_venta=?, costo=?, stock=?, unidad=?, fecha_vencimiento=?, stock_minimo=?, categoria=?
            WHERE id=?
        """, (nombre, precio, costo, stock, unidad, fecha_vencimiento, stock_minimo, categoria, id_p))
        conexion.commit()
        return True
    except sqlite3.Error as e:
        print(f"Error al actualizar: {e}")
        return False
    finally:
        conexion.close()

def eliminar_producto(id_recibido):
    # Usar función de conexión (ejemplo: conectar_db o get_db)
    conexion = conectar() 
    cursor = conexion.cursor()
    
    cursor.execute("DELETE FROM productos WHERE id = ?", (id_recibido,))
    
    conexion.commit()
    conexion.close()

def guardar_cierre_db(datos):
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO cierres (fecha, efectivo, tarjeta, otros, fiados, total, turno)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (
        datos['fecha'], datos['efectivo'], datos['tarjeta'], 
        datos['otros'], datos['fiados'], datos['total'], 'Único'
    ))
    conn.commit()
    conn.close()

def obtener_historial_db():
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM cierres ORDER BY id DESC')
    # Convertimos a lista de diccionarios para que Flask lo lea fácil
    columnas = [column[0] for column in cursor.description]
    resultados = [dict(zip(columnas, fila)) for fila in cursor.fetchall()]
    conn.close()
    return resultados

def guardar_fiado_db(nombre, monto, detalle):
    from datetime import datetime
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO fiados (nombre_cliente, monto_total, monto_pagado, fecha, estado, detalle)
        VALUES (?, ?, 0, ?, 'pendiente', ?)
    ''', (nombre, monto, datetime.now().strftime('%d-%m-%Y %H:%M'), detalle))
    conn.commit()
    conn.close()

def obtener_fiados_db():
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM fiados ORDER BY id DESC")
    columnas = [c[0] for c in cursor.description]
    resultados = [dict(zip(columnas, fila)) for fila in cursor.fetchall()]
    conn.close()
    return resultados

def saldar_fiado_db(fiado_id, monto_pago):
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("SELECT monto_total, monto_pagado FROM fiados WHERE id = ?", (fiado_id,))
    fiado = cursor.fetchone()
    if not fiado:
        conn.close()
        return False, "Fiado no encontrado"
    
    nuevo_pagado = fiado['monto_pagado'] + monto_pago
    estado = 'pagado' if nuevo_pagado >= fiado['monto_total'] else 'parcial'
    
    cursor.execute("""
        UPDATE fiados SET monto_pagado = ?, estado = ? WHERE id = ?
    """, (nuevo_pagado, estado, fiado_id))
    conn.commit()
    conn.close()
    return True, estado

def verificar_usuario(username, password):
    import hashlib
    conn = conectar()
    cursor = conn.cursor()
    password_hash = hashlib.sha256(password.encode()).hexdigest()
    cursor.execute("""
        SELECT id, username, nombre, rol 
        FROM usuarios 
        WHERE username = ? AND password = ? AND activo = 1
    """, (username, password_hash))
    usuario = cursor.fetchone()
    conn.close()
    if usuario:
        return dict(usuario)
    return None

def crear_usuario(username, password, nombre, rol='empleado'):
    import hashlib
    conn = conectar()
    cursor = conn.cursor()
    try:
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        cursor.execute("""
            INSERT INTO usuarios (username, password, nombre, rol)
            VALUES (?, ?, ?, ?)
        """, (username, password_hash, nombre, rol))
        conn.commit()
        return True, "Usuario creado"
    except Exception as e:
        return False, str(e)
    finally:
        conn.close()

def obtener_usuarios():
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("SELECT id, username, nombre, rol, activo FROM usuarios ORDER BY id")
    columnas = [c[0] for c in cursor.description]
    resultados = [dict(zip(columnas, fila)) for fila in cursor.fetchall()]
    conn.close()
    return resultados

# ── FACTURAS ──
def guardar_factura_db(datos):
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO facturas (numero_factura, proveedor, rut_proveedor, fecha, monto_total, productos, estado)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (datos['numero_factura'], datos['proveedor'], datos['rut_proveedor'],
          datos['fecha'], datos['monto_total'], datos['productos'], datos['estado']))
    conn.commit()
    conn.close()

def obtener_facturas_db():
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM facturas ORDER BY id DESC')
    columnas = [c[0] for c in cursor.description]
    resultados = [dict(zip(columnas, fila)) for fila in cursor.fetchall()]
    conn.close()
    return resultados

def actualizar_estado_factura_db(factura_id, estado):
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("UPDATE facturas SET estado = ? WHERE id = ?", (estado, factura_id))
    conn.commit()
    conn.close()

# ── PRODUCTOS MUERTOS ──
def obtener_productos_muertos_db(dias=60):
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(f"""
        SELECT p.id, p.nombre, p.codigo_barra, p.stock,
               MAX(m.fecha) as ultima_venta
        FROM productos p
        LEFT JOIN movimientos_stock m 
            ON p.id = m.producto_id AND m.tipo = 'VENTA'
        WHERE p.activo = 1 AND p.stock > 0
        GROUP BY p.id
        HAVING ultima_venta IS NULL 
            OR ultima_venta < datetime('now', '-{dias} days')
        ORDER BY ultima_venta ASC
    """)
    columnas = [c[0] for c in cursor.description]
    resultados = [dict(zip(columnas, fila)) for fila in cursor.fetchall()]
    conn.close()
    return resultados

# ── CONFIGURACIÓN ──
def obtener_config_db():
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute('SELECT clave, valor FROM configuracion')
    config = {row[0]: row[1] for row in cursor.fetchall()}
    conn.close()
    return config

def guardar_config_db(datos):
    conn = conectar()
    cursor = conn.cursor()
    for clave, valor in datos.items():
        cursor.execute("""
            INSERT OR REPLACE INTO configuracion (clave, valor) 
            VALUES (?, ?)
        """, (clave, valor))
    conn.commit()
    conn.close()

def cambiar_password_db(usuario_id, password_actual, password_nueva):
    import hashlib
    conn = conectar()
    cursor = conn.cursor()
    
    # Verificar contraseña actual
    hash_actual = hashlib.sha256(password_actual.encode()).hexdigest()
    cursor.execute("SELECT id FROM usuarios WHERE id = ? AND password = ?", 
                   (usuario_id, hash_actual))
    
    if not cursor.fetchone():
        conn.close()
        return False, "Contraseña actual incorrecta"
    
    # Actualizar contraseña
    hash_nueva = hashlib.sha256(password_nueva.encode()).hexdigest()
    cursor.execute("UPDATE usuarios SET password = ? WHERE id = ?", 
                   (hash_nueva, usuario_id))
    conn.commit()
    conn.close()
    return True, "Contraseña actualizada"    

def generar_reporte_excel(datos_cierre):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import os
    from datetime import datetime

    # Crear carpeta reportes si no existe
    carpeta = os.path.join(get_app_dir(), 'reportes')
    os.makedirs(carpeta, exist_ok=True)

    # Nombre del archivo con fecha y hora
    fecha_archivo = datetime.now().strftime('%Y-%m-%d_%H-%M')
    ruta = os.path.join(carpeta, f'cierre_{fecha_archivo}.xlsx')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cierre de Caja"

    # Estilos
    estilo_titulo  = Font(bold=True, size=14)
    estilo_header  = Font(bold=True, color="FFFFFF")
    fill_header    = PatternFill("solid", fgColor="1E293B")
    fill_total     = PatternFill("solid", fgColor="DCFCE7")
    centro         = Alignment(horizontal="center")

    # ── ENCABEZADO ──
    ws.merge_cells('A1:D1')
    ws['A1'] = datos_cierre.get('nombre_negocio', 'EL HISTORICO')
    ws['A1'].font      = estilo_titulo
    ws['A1'].alignment = centro

    ws.merge_cells('A2:D2')
    ws['A2'] = f"Cierre de Caja — {datos_cierre['fecha']}"
    ws['A2'].alignment = centro

    ws.append([])  # fila vacía

    # ── RESUMEN DEL TURNO ──
    headers = ['Concepto', 'Monto']
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font      = estilo_header
        cell.fill      = fill_header
        cell.alignment = centro

    filas_resumen = [
        ('Efectivo',  datos_cierre.get('efectivo', 0)),
        ('Tarjeta',   datos_cierre.get('tarjeta', 0)),
        ('Otros',     datos_cierre.get('otros', 0)),
        ('Fiados',    datos_cierre.get('fiados', 0)),
        ('TOTAL CAJA',datos_cierre.get('total', 0)),
    ]

    for i, (concepto, monto) in enumerate(filas_resumen):
        ws.append([concepto, f'${monto:,.0f}'])
        if concepto == 'TOTAL CAJA':
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
                cell.fill = fill_total

    ws.append([])  # fila vacía

    # ── DETALLE DE VENTAS DEL DÍA ──
    ws.append(['DETALLE DE VENTAS'])
    ws[ws.max_row][0].font = Font(bold=True, size=12)

    headers_ventas = ['N° Venta', 'Hora', 'Método Pago', 'Total']
    ws.append(headers_ventas)
    for cell in ws[ws.max_row]:
        cell.font      = estilo_header
        cell.fill      = fill_header
        cell.alignment = centro

    # Traer ventas del día desde la DB
    conn = conectar()
    cursor = conn.cursor()
    from datetime import datetime as dt
    hoy = dt.now().strftime('%Y-%m-%d')
    cursor.execute("""
        SELECT id, fecha, metodo_pago, total 
        FROM ventas 
        WHERE DATE(fecha) = ?
        ORDER BY fecha DESC
    """, (hoy,))
    ventas = cursor.fetchall()
    conn.close()

    for venta in ventas:
        hora = venta['fecha'].split(' ')[1][:5] if venta['fecha'] else '--:--'
        ws.append([
            f'#{venta["id"]}',
            hora,
            venta['metodo_pago'],
            f'${venta["total"]:,.0f}'
        ])

    # ── AJUSTAR ANCHOS ──
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

    wb.save(ruta)
    print(f"✅ Reporte guardado en: {ruta}")

def generar_excel_dia(fecha):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import os

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas del Dia"

    estilo_header = Font(bold=True, color="FFFFFF")
    fill_header   = PatternFill("solid", fgColor="1E293B")
    fill_total    = PatternFill("solid", fgColor="DCFCE7")
    centro        = Alignment(horizontal="center")

    # Encabezado
    ws.merge_cells('A1:E1')
    ws['A1'] = f"Reporte de Ventas — {fecha}"
    ws['A1'].font      = Font(bold=True, size=13)
    ws['A1'].alignment = centro
    ws.append([])

    # Headers tabla
    ws.append(['N° Venta', 'Hora', 'Producto', 'Cantidad', 'Subtotal', 'Método Pago', 'Total Venta'])
    for cell in ws[ws.max_row]:
        cell.font      = estilo_header
        cell.fill      = fill_header
        cell.alignment = centro

    # Traer ventas del día con detalle
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT v.id, v.fecha, v.total, v.metodo_pago,
               p.nombre, dv.cantidad, dv.subtotal
        FROM ventas v
        JOIN detalle_venta dv ON dv.venta_id = v.id
        JOIN productos p ON p.id = dv.producto_id
        WHERE DATE(v.fecha) = ?
        ORDER BY v.fecha DESC
    """, (fecha,))
    filas = cursor.fetchall()

    # Totales por método
    cursor.execute("""
        SELECT metodo_pago, SUM(total) as total
        FROM ventas
        WHERE DATE(fecha) = ?
        GROUP BY metodo_pago
    """, (fecha,))
    totales = cursor.fetchall()
    conn.close()

    total_dia = 0
    for fila in filas:
        hora = fila['fecha'].split(' ')[1][:5] if fila['fecha'] else '--:--'
        ws.append([
            f'#{fila["id"]}',
            hora,
            fila['nombre'],
            fila['cantidad'],
            f'${fila["subtotal"]:,.0f}',
            fila['metodo_pago'],
            f'${fila["total"]:,.0f}'
        ])

    ws.append([])

    # Resumen
    ws.append(['RESUMEN DEL DÍA'])
    ws[ws.max_row][0].font = Font(bold=True, size=11)

    for t in totales:
        total_dia += t['total']
        ws.append([t['metodo_pago'].capitalize(), f'${t["total"]:,.0f}'])

    ws.append(['TOTAL', f'${total_dia:,.0f}'])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.fill = fill_total

    # Anchos
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14

    # Guardar en memoria para descargar
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# Feature 5: anular venta
def anular_venta_db(venta_id):
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("SELECT anulada FROM ventas WHERE id = ?", (venta_id,))
    v = cursor.fetchone()
    if not v or v['anulada']:
        conn.close()
        return False, "Venta no encontrada o ya anulada"
    cursor.execute("SELECT producto_id, cantidad FROM detalle_venta WHERE venta_id = ?", (venta_id,))
    items = cursor.fetchall()
    for item in items:
        cursor.execute("UPDATE productos SET stock = stock + ? WHERE id = ?", (item['cantidad'], item['producto_id']))
        cursor.execute("INSERT INTO movimientos_stock (producto_id, tipo, cantidad, motivo) VALUES (?, 'ANULACION', ?, 'Anulación venta #' || ?)",
                       (item['producto_id'], item['cantidad'], venta_id))
    cursor.execute("UPDATE ventas SET anulada = 1 WHERE id = ?", (venta_id,))
    conn.commit()
    conn.close()
    return True, "Venta anulada"


# Feature 6: historial de movimientos de stock
def obtener_historial_stock_db(limite=100):
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT m.id, m.tipo, m.cantidad, m.fecha, m.motivo,
               p.nombre as producto_nombre, p.codigo_barra
        FROM movimientos_stock m
        JOIN productos p ON p.id = m.producto_id
        ORDER BY m.id DESC
        LIMIT ?
    """, (limite,))
    columnas = [c[0] for c in cursor.description]
    resultados = [dict(zip(columnas, fila)) for fila in cursor.fetchall()]
    conn.close()
    return resultados


def ajuste_manual_stock_db(producto_id, cantidad, motivo='Ajuste manual'):
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("SELECT stock, nombre FROM productos WHERE id = ?", (producto_id,))
    p = cursor.fetchone()
    if not p:
        conn.close()
        return False, "Producto no encontrado"
    nuevo_stock = p['stock'] + cantidad
    if nuevo_stock < 0:
        conn.close()
        return False, f"Stock insuficiente. Stock actual: {p['stock']}"
    cursor.execute("UPDATE productos SET stock = ? WHERE id = ?", (nuevo_stock, producto_id))
    tipo = 'ENTRADA' if cantidad > 0 else 'SALIDA'
    cursor.execute("INSERT INTO movimientos_stock (producto_id, tipo, cantidad, motivo) VALUES (?, ?, ?, ?)",
                   (producto_id, tipo, abs(cantidad), motivo))
    conn.commit()
    conn.close()
    return True, f"Stock actualizado: {nuevo_stock}"


# Feature 7: proveedores
def obtener_proveedores_db():
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM proveedores WHERE activo = 1 ORDER BY nombre")
    columnas = [c[0] for c in cursor.description]
    resultados = [dict(zip(columnas, fila)) for fila in cursor.fetchall()]
    conn.close()
    return resultados


def guardar_proveedor_db(datos):
    conn = conectar()
    cursor = conn.cursor()
    if datos.get('id'):
        cursor.execute("""
            UPDATE proveedores SET nombre=?, rut=?, telefono=?, email=?, notas=?
            WHERE id=?
        """, (datos['nombre'], datos.get('rut', ''), datos.get('telefono', ''),
              datos.get('email', ''), datos.get('notas', ''), datos['id']))
    else:
        cursor.execute("""
            INSERT INTO proveedores (nombre, rut, telefono, email, notas)
            VALUES (?, ?, ?, ?, ?)
        """, (datos['nombre'], datos.get('rut', ''), datos.get('telefono', ''),
              datos.get('email', ''), datos.get('notas', '')))
    conn.commit()
    conn.close()


def eliminar_proveedor_db(proveedor_id):
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("UPDATE proveedores SET activo = 0 WHERE id = ?", (proveedor_id,))
    conn.commit()
    conn.close()


def obtener_productos_por_vencer(dias=7):
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, nombre, codigo_barra, stock, fecha_vencimiento
        FROM productos
        WHERE fecha_vencimiento IS NOT NULL
        AND fecha_vencimiento != ''
        AND DATE(fecha_vencimiento) <= DATE('now', '+' || ? || ' days')
        AND DATE(fecha_vencimiento) >= DATE('now')
        AND activo = 1
        ORDER BY fecha_vencimiento ASC
    """, (dias,))
    columnas = [c[0] for c in cursor.description]
    resultados = [dict(zip(columnas, fila)) for fila in cursor.fetchall()]
    conn.close()
    return resultados


