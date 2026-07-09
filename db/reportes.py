import os
from datetime import datetime
from .conexion import conectar, get_app_dir


def generar_reporte_excel(datos_cierre):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    # Crear carpeta reportes si no existe
    carpeta = os.path.join(get_app_dir(), 'reportes')
    os.makedirs(carpeta, exist_ok=True)

    # Nombre del archivo con fecha y hora
    fecha_archivo = datetime.now().strftime('%Y-%m-%d_%H-%M')
    ruta = os.path.join(carpeta, f'cierre_{fecha_archivo}.xlsx')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cierre de Caja"

    # Estilos
    estilo_titulo  = Font(bold=True, size=14)
    estilo_header  = Font(bold=True, color="FFFFFF")
    fill_header    = PatternFill("solid", fgColor="1E293B")
    fill_total     = PatternFill("solid", fgColor="DCFCE7")
    centro         = Alignment(horizontal="center")

    # ── ENCABEZADO ──
    ws.merge_cells('A1:D1')
    ws['A1'] = datos_cierre.get('nombre_negocio', 'EL HISTORICO')
    ws['A1'].font      = estilo_titulo
    ws['A1'].alignment = centro

    ws.merge_cells('A2:D2')
    ws['A2'] = f"Cierre de Caja — {datos_cierre['fecha']}"
    ws['A2'].alignment = centro

    ws.append([])  # fila vacía

    # ── RESUMEN DEL TURNO ──
    headers = ['Concepto', 'Monto']
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font      = estilo_header
        cell.fill      = fill_header
        cell.alignment = centro

    filas_resumen = [
        ('Efectivo',  datos_cierre.get('efectivo', 0)),
        ('Tarjeta',   datos_cierre.get('tarjeta', 0)),
        ('Otros',     datos_cierre.get('otros', 0)),
        ('Fiados',    datos_cierre.get('fiados', 0)),
        ('TOTAL CAJA',datos_cierre.get('total', 0)),
    ]

    for i, (concepto, monto) in enumerate(filas_resumen):
        ws.append([concepto, f'${monto:,.0f}'])
        if concepto == 'TOTAL CAJA':
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
                cell.fill = fill_total

    ws.append([])  # fila vacía

    # ── DETALLE DE VENTAS DEL DÍA ──
    ws.append(['DETALLE DE VENTAS'])
    ws[ws.max_row][0].font = Font(bold=True, size=12)

    headers_ventas = ['N° Venta', 'Hora', 'Método Pago', 'Total']
    ws.append(headers_ventas)
    for cell in ws[ws.max_row]:
        cell.font      = estilo_header
        cell.fill      = fill_header
        cell.alignment = centro

    # Traer ventas del día desde la DB
    conn = conectar()
    cursor = conn.cursor()
    hoy = datetime.now().strftime('%Y-%m-%d')
    cursor.execute("""
        SELECT id, fecha, metodo_pago, total
        FROM ventas
        WHERE DATE(fecha) = ? AND anulada = 0
        ORDER BY fecha DESC
    """, (hoy,))
    ventas = cursor.fetchall()
    conn.close()

    for venta in ventas:
        hora = venta['fecha'].split(' ')[1][:5] if venta['fecha'] else '--:--'
        ws.append([
            f'#{venta["id"]}',
            hora,
            venta['metodo_pago'],
            f'${venta["total"]:,.0f}'
        ])

    # ── AJUSTAR ANCHOS ──
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

    wb.save(ruta)
    print(f"Reporte guardado en: {ruta}")

def generar_excel_dia(fecha):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas del Dia"

    estilo_header = Font(bold=True, color="FFFFFF")
    fill_header   = PatternFill("solid", fgColor="1E293B")
    fill_total    = PatternFill("solid", fgColor="DCFCE7")
    centro        = Alignment(horizontal="center")

    # Encabezado
    ws.merge_cells('A1:E1')
    ws['A1'] = f"Reporte de Ventas — {fecha}"
    ws['A1'].font      = Font(bold=True, size=13)
    ws['A1'].alignment = centro
    ws.append([])

    # Headers tabla
    ws.append(['N° Venta', 'Hora', 'Producto', 'Cantidad', 'Subtotal', 'Método Pago', 'Total Venta'])
    for cell in ws[ws.max_row]:
        cell.font      = estilo_header
        cell.fill      = fill_header
        cell.alignment = centro

    # Traer ventas del día con detalle
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT v.id, v.fecha, v.total, v.metodo_pago,
               p.nombre, dv.cantidad, dv.subtotal
        FROM ventas v
        JOIN detalle_venta dv ON dv.venta_id = v.id
        JOIN productos p ON p.id = dv.producto_id
        WHERE DATE(v.fecha) = ? AND v.anulada = 0
        ORDER BY v.fecha DESC
    """, (fecha,))
    filas = cursor.fetchall()

    # Totales por método
    cursor.execute("""
        SELECT metodo_pago, SUM(total) as total
        FROM ventas
        WHERE DATE(fecha) = ? AND anulada = 0
        GROUP BY metodo_pago
    """, (fecha,))
    totales = cursor.fetchall()
    conn.close()

    total_dia = 0
    for fila in filas:
        hora = fila['fecha'].split(' ')[1][:5] if fila['fecha'] else '--:--'
        ws.append([
            f'#{fila["id"]}',
            hora,
            fila['nombre'],
            fila['cantidad'],
            f'${fila["subtotal"]:,.0f}',
            fila['metodo_pago'],
            f'${fila["total"]:,.0f}'
        ])

    ws.append([])

    # Resumen
    ws.append(['RESUMEN DEL DÍA'])
    ws[ws.max_row][0].font = Font(bold=True, size=11)

    for t in totales:
        total_dia += t['total']
        ws.append([t['metodo_pago'].capitalize(), f'${t["total"]:,.0f}'])

    ws.append(['TOTAL', f'${total_dia:,.0f}'])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.fill = fill_total

    # Anchos
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14

    # Guardar en memoria para descargar
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def obtener_datos_graficos():
    conn = conectar()
    cursor = conn.cursor()

    # 1. Efectivo vs Tarjeta vs Otros (últimos 30 días) — sin ventas anuladas
    cursor.execute("""
        SELECT metodo_pago, SUM(total) as total
        FROM ventas
        WHERE fecha >= datetime('now', '-30 days') AND anulada = 0
        GROUP BY metodo_pago
    """)
    metodos = {row['metodo_pago']: row['total'] for row in cursor.fetchall()}

    # 2. Productos más vendidos (top 8) — sin ventas anuladas
    cursor.execute("""
        SELECT p.nombre, SUM(dv.cantidad) as total_vendido
        FROM detalle_venta dv
        JOIN productos p ON p.id = dv.producto_id
        JOIN ventas v ON v.id = dv.venta_id
        WHERE v.anulada = 0
        GROUP BY p.id
        ORDER BY total_vendido DESC
        LIMIT 8
    """)
    productos = [{'nombre': r['nombre'], 'cantidad': r['total_vendido']}
                 for r in cursor.fetchall()]

    # 3. Evolución por semana (últimas 8 semanas) — sin ventas anuladas
    cursor.execute("""
        SELECT
            strftime('%Y-W%W', fecha) as semana,
            SUM(total) as total
        FROM ventas
        WHERE fecha >= datetime('now', '-56 days') AND anulada = 0
        GROUP BY semana
        ORDER BY semana ASC
    """)
    semanas = [{'semana': r['semana'], 'total': r['total']}
               for r in cursor.fetchall()]

    conn.close()
    return {
        'metodos': metodos,
        'productos': productos,
        'semanas': semanas
    }
